import os
import random
from datetime import datetime
from argparse import ArgumentParser
from signal import signal, alarm, SIGALRM
from tabulate import tabulate
from openpyxl import load_workbook


def imprimir_tarjetas():

    parser = ArgumentParser(description="Argumento para decidir entre crear tarjetas o jugar al bingo")
    parser.add_argument("--tarjeta", "-t", action="store_true", help="Crea 6 tarjetas de bingo aleatorias listas para imprimirse")
    args = parser.parse_args()

    return args.tarjeta


def crear_tarjetas():

    tarjetas = ['1','2','3','4','5','6']
    letras = ['B','I','N','G','O']
    offset = {'B':0,'I':15,'N':30,'G':45,'O':60}
    posiciones = ['1','2','3','4','5']
    correspondencia = list()

    for t in tarjetas:
        bolitas = [b for b in range(1,76)]
        for l in letras:
            for p in posiciones:
                n = offset[l]
                while True:
                    b = random.randint(1 + n ,15 + n)   
                    if b in bolitas:
                        bolitas.remove(b)
                        pareja = (f'{t}{l}{p}',b)
                        correspondencia.append(pareja)
                        break

    wb = load_workbook('tarjetas/plantilla_tarjetas.xlsx')
    ws = wb['Sheet1']

    for pareja in correspondencia:
        for x in range(1,34):
            for y in range(1,15):
                if pareja[0] == ws.cell(row=x,column=y).value:
                    ws.cell(row=x,column=y).value = pareja[1]

    archivo = f'tarjetas/T{datetime.now().strftime("%d%m%Y%H%M%S")}.xlsx'
    wb.save(archivo)
    respuesta = (f'Archivo guardado como {archivo}!')

    return respuesta


def juego_nuevo():

    b_list = ['B']
    i_list = ['I']
    n_list = ['N']
    g_list = ['G']
    o_list = ['O']

    b_dict = {
        'B':{'offset':0,'lista':b_list},
        'I':{'offset':15,'lista':i_list},
        'N':{'offset':30,'lista':n_list},
        'G':{'offset':45,'lista':g_list},
        'O':{'offset':60,'lista':o_list},
    }

    bolitas = list()

    for l in b_dict.keys():

        n = b_dict[l]['offset']
        a = b_dict[l]['lista']

        for x in range(1 + n ,16 + n):

            b = f'{l}{str(x)}'
            bolitas.append(b)
            a.append(f'{l}{str(x)}')

    tablero = [b_list,i_list,n_list,g_list,o_list]
    random.shuffle(bolitas)
    juego = {'bolitas':bolitas,'tablero':tablero}

    return juego


def marcar_bolita_en_tablero(bolita,tablero):

        for line in tablero:
            for i in range(1,16):
                if line[i] == bolita:
                    line[i] = bolita+'*'


def reproducir_audio(bolita):
    
    print(bolita)
    print('Reproduciendo audio...')
    audio = f"{bolita}.aifc"
    if audio in os.listdir('audios'):
        os.system(f"afplay audios/{audio}")
    else:
        print(f'Error: Archivo audios/{audio}')
    

def mostrar_tablero(juego,turno):

    signal(SIGALRM, lambda y:1/0)
    
    try:
        alarm(3)
        x = input()
        if x or x == '':
            alarm(0)
            print(tabulate(juego['tablero']))
            print('Bolitas sacadas:',turno)
            print('Bolitas pendientes:',len(juego['bolitas']) - turno)
            continuar = input('Escribe q para terminar o cualquier otra tecla si el juego continua: ')
            if continuar == 'q':
                print('Felicidades al ganador... Adios!')
                os.system("afplay audios/felicidades.aifc")
                exit(1)
            else:
                print('Continuamos!')

    except TypeError:
        print("Siguiente bolita...")


def tablero_final(juego):

    print('Se terminaron las bolitas, revisen sus tarjetas!')
    print(tabulate(juego['tablero']))
    input()
    print('Adios!')
    exit(1)
    
def iniciar_juego(juego):

    print('Empenzando el juego!')
    print('Presiona enter para hacer pausa y mostrar el tablero\n')

    turno = 1
    for bolita in juego['bolitas']:
        if turno == 75:
            tablero_final(juego)
        else:
            reproducir_audio(bolita)
            marcar_bolita_en_tablero(bolita,juego['tablero'])
            mostrar_tablero(juego,turno)
            turno += 1


if __name__ == '__main__':

    if imprimir_tarjetas():
        print(crear_tarjetas())
    else:
        juego = juego_nuevo()
        iniciar_juego(juego)
